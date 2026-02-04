import threading
import time
import os
from flask import Flask, render_template, jsonify, request
from datetime import datetime
import urllib3
from flask_cors import CORS  # Add this
from pathlib import Path
from openpyxl import load_workbook


urllib3.disable_warnings()
app = Flask(__name__)
CORS(app)  # Enable CORS for all routes

# Global shared state
CHECK_INTERVAL = 15 * 60
MAX_RETRIES = 3

monitoring_results = {
    'total': 0,
    'checked': 0,
    'failed': [],
    'last_check': None,
    'is_running': False,
    'retry_in_progress': False
}

results_lock = threading.Lock()

def load_websites_from_excel():
    """Load websites from Excel using openpyxl (no pandas) - Optimized for Render.com"""
    try:
        # Get the directory of the current script
        current_dir = Path(__file__).parent.resolve()

        # Define possible paths (prioritize different locations)
        possible_paths = [
            current_dir / 'Adani-BUWise-Websites.xlsx',  # Same directory as script
            current_dir / 'data' / 'Adani-BUWise-Websites.xlsx',  # data/ subfolder
            Path('/app/Adani-BUWise-Websites.xlsx'),  # Render common mount
            Path('Adani-BUWise-Websites.xlsx'),  # Current working dir
            Path('/data/Adani-BUWise-Websites.xlsx'),  # Persistent disk
        ]

        excel_file = None
        used_path = None

        for path in possible_paths:
            if path.exists():
                excel_file = path
                used_path = path
                print(f"✓ Found Excel file at: {path}")
                break

        if excel_file is None:
            print("⚠ Excel file not found in any location, using demo data")
            return get_demo_websites()

        # Load workbook using openpyxl (lightweight, no pandas needed)
        # read_only=True is faster and uses less memory
        wb = load_workbook(filename=str(excel_file), read_only=True, data_only=True)

        # Get the first sheet (or specify sheet name: wb['Sheet1'])
        ws = wb.active

        websites = []

        # Get header row to find column indices
        headers = {}
        header_row = next(ws.iter_rows(values_only=True))

        for idx, cell_value in enumerate(header_row):
            if cell_value:
                headers[str(cell_value).strip().upper()] = idx

        # Check required columns exist
        if 'BU' not in headers or 'WEBSITES' not in headers:
            print("⚠ Required columns 'BU' or 'Websites' not found")
            wb.close()
            return get_demo_websites()

        bu_col = headers['BU']
        websites_col = headers['WEBSITES']

        # Iterate through data rows (skip header)
        for row in ws.iter_rows(min_row=2, values_only=True):
            try:
                bu = str(row[bu_col]).strip() if row[bu_col] else ''
                cell = str(row[websites_col]).strip() if row[websites_col] else ''

                # Skip empty rows
                if not cell or cell.lower() in ['nan', 'none', '']:
                    continue

                # Normalize line endings and split by newlines or commas
                cell = cell.replace('\r\n', '\n').replace('\r', '\n')
                raw_urls = []

                for part in cell.split('\n'):
                    raw_urls.extend([u.strip() for u in part.split(',') if u.strip()])

                # Process each URL
                for url in raw_urls:
                    if not url or url.lower() in ['nan', 'none']:
                        continue

                    # Ensure proper URL format
                    if not url.startswith(('http://', 'https://')):
                        url = 'https://' + url
                    url = url.replace(' ', '').rstrip('/')

                    # Clean name
                    name = url.replace('https://', '').replace('http://', '').replace('www.', '')

                    websites.append({
                        'bu': bu,
                        'url': url,
                        'name': name
                    })

            except Exception as row_error:
                print(f"⚠ Error processing row: {row_error}")
                continue

        # Close workbook to free memory
        wb.close()

        print(f"✓ Successfully loaded {len(websites)} websites from {used_path}")
        return websites

    except Exception as e:
        print(f"✗ Error reading Excel: {e}")
        import traceback
        traceback.print_exc()
        return get_demo_websites()


def get_demo_websites():
    """Return demo websites when Excel is not available"""
    return [
        {
            'bu': 'Demo BU',
            'url': 'https://example.com',
            'name': 'example.com'
        }
    ]


# Global cache to track which sites need Selenium
selenium_required = set()


def check_website(site_info):
    import requests
    import urllib3
    import random
    import time
    from datetime import datetime
    from requests.adapters import HTTPAdapter
    from urllib3.util.retry import Retry

    urllib3.disable_warnings()
    url = site_info['url']

    # Check if this site previously needed Selenium
    if url in selenium_required:
        return check_website_selenium(site_info)

    # Try lightweight request first
    user_agents = [
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
        'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.2 Safari/605.1.15'
    ]

    def get_headers():
        return {
            'User-Agent': random.choice(user_agents),
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
            'Accept-Language': 'en-US,en;q=0.9',
            'Accept-Encoding': 'gzip, deflate',
            'Connection': 'keep-alive',
            'Upgrade-Insecure-Requests': '1'
        }

    session = requests.Session()
    retry = Retry(total=2, backoff_factor=0.5)
    session.mount('https://', HTTPAdapter(max_retries=retry))

    try:
        # Quick HEAD check
        response = session.head(url, headers=get_headers(), timeout=8, verify=False, allow_redirects=True)

        if response.status_code == 403:
            raise requests.exceptions.RequestException("403 Forbidden")

        if response.status_code == 405:
            response = session.get(url, headers=get_headers(), timeout=10, verify=False, stream=True)
            response.close()

        success = 200 <= response.status_code < 400

        if success:
            return {
                'success': True,
                'status_code': response.status_code,
                'url': url,
                'bu': site_info['bu'],
                'name': site_info['name'],
                'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'error': None
            }
        else:
            raise requests.exceptions.RequestException(f"Status {response.status_code}")

    except Exception as e:
        # If requests fail, try Selenium and mark for future
        print(f"  Requests failed for {url}, trying Selenium...")
        selenium_required.add(url)
        return check_website_selenium(site_info)

    finally:
        session.close()


def check_website_selenium(site_info):
    """Selenium fallback for stubborn sites"""
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.chrome.service import Service
    from datetime import datetime
    import logging

    logging.getLogger('selenium').setLevel(logging.ERROR)

    url = site_info['url']
    driver = None

    try:
        chrome_options = Options()
        chrome_options.add_argument('--headless')
        chrome_options.add_argument('--no-sandbox')
        chrome_options.add_argument('--disable-dev-shm-usage')
        chrome_options.add_argument('--disable-gpu')
        chrome_options.add_argument('--single-process')
        chrome_options.add_argument('--disable-extensions')
        chrome_options.add_argument('--window-size=1920,1080')
        chrome_options.add_argument('--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36')

        # Try to use existing ChromeDriver or download
        try:
            from webdriver_manager.chrome import ChromeDriverManager
            service = Service(ChromeDriverManager().install())
        except:
            service = Service('/usr/bin/chromedriver')  # Common Render path

        driver = webdriver.Chrome(service=service, options=chrome_options)
        driver.set_page_load_timeout(15)

        driver.get(url)

        # Simple check - if page loaded and has content
        success = driver.find_element("tag name", "body") is not None

        return {
            'success': success,
            'status_code': 200 if success else 403,
            'url': url,
            'bu': site_info['bu'],
            'name': site_info['name'],
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'error': None if success else 'Selenium check failed',
            'method': 'selenium'  # Track which method worked
        }

    except Exception as e:
        return {
            'success': False,
            'status_code': 0,
            'url': url,
            'bu': site_info['bu'],
            'name': site_info['name'],
            'error': str(e)[:50],
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }

    finally:
        if driver:
            try:
                driver.quit()
            except:
                pass


def get_demo_websites():
    return [{'bu': 'Demo', 'url': 'https://www.google.com', 'name': 'google.com'}]


def monitor_websites():
    """Main monitoring loop"""
    global monitoring_results

    monitoring_results['is_running'] = True

    while monitoring_results['is_running']:
        websites = load_websites_from_excel()

        with results_lock:
            monitoring_results['total'] = len(websites)
            monitoring_results['checked'] = 0

        print(f"\n🔍 Checking {len(websites)} websites...")

        for i, site in enumerate(websites, start=1):
            if not monitoring_results['is_running']:
                break

            result = check_website(site)

            with results_lock:
                monitoring_results['checked'] = i

                if not result['success']:
                    existing = next((f for f in monitoring_results['failed'] if f['url'] == result['url']), None)
                    if not existing:
                        result['retry_count'] = 0
                        monitoring_results['failed'].append(result)
                else:
                    # Remove from failed if recovered
                    monitoring_results['failed'] = [f for f in monitoring_results['failed'] if f['url'] != result['url']]

            time.sleep(0.5)

        with results_lock:
            monitoring_results['last_check'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        print(f"✅ Cycle done. Failed: {len(monitoring_results['failed'])}")

        # Sleep 15 minutes
        for _ in range(CHECK_INTERVAL):
            if not monitoring_results['is_running']:
                break
            time.sleep(1)

    print("🛑 Monitoring stopped")


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/api/start', methods=['POST'])
def start_monitoring():
    if not monitoring_results['is_running']:
        t = threading.Thread(target=monitor_websites, daemon=True)
        t.start()
        return jsonify({'status': 'started'})
    return jsonify({'status': 'already_running'})


@app.route('/api/stop', methods=['POST'])
def stop_monitoring():
    monitoring_results['is_running'] = False
    return jsonify({'status': 'stopped'})


@app.route('/api/status')
def status():
    """Return current status - same for all users"""
    with results_lock:
        return jsonify({
            'total': monitoring_results['total'],
            'checked': monitoring_results['checked'],
            'failed': [f.copy() for f in monitoring_results['failed']],
            'last_check': monitoring_results['last_check'],
            'is_running': monitoring_results['is_running']
        })


@app.route('/api/retry', methods=['POST'])
def retry_website():
    """Retry single website"""
    global monitoring_results

    data = request.get_json()
    if not data:
        return jsonify({'success': False, 'error': 'No JSON data'}), 400

    url = data.get('url')
    if not url:
        return jsonify({'success': False, 'error': 'No URL provided'}), 400

    with results_lock:
        site_index = None
        site_info = None

        for i, site in enumerate(monitoring_results['failed']):
            if site['url'] == url:
                site_index = i
                site_info = site
                break

        if site_index is None:
            return jsonify({'success': False, 'error': 'Site not found'}), 404

        retry_count = site_info.get('retry_count', 0)
        if retry_count >= MAX_RETRIES:
            return jsonify({
                'success': False,
                'error': f'Max retries ({MAX_RETRIES}) reached',
                'retry_count': retry_count
            }), 429

        monitoring_results['retry_in_progress'] = True

    # Perform retry outside lock
    print(f"🔄 Retrying: {url} (attempt {retry_count + 1}/{MAX_RETRIES})")
    result = check_website(site_info)

    with results_lock:
        monitoring_results['retry_in_progress'] = False

        if result['success']:
            monitoring_results['failed'].pop(site_index)
            print(f"   ✅ Success! Removed from failed list.")
            return jsonify({
                'success': True,
                'message': 'Website is accessible',
                'failed_count': len(monitoring_results['failed'])
            })
        else:
            site_info['retry_count'] = retry_count + 1
            site_info['last_retry'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            site_info['last_error'] = result.get('error', 'Unknown')
            print(f"   ❌ Failed. Count: {site_info['retry_count']}")
            return jsonify({
                'success': False,
                'error': result.get('error', 'Check failed'),
                'retry_count': site_info['retry_count'],
                'max_retries': MAX_RETRIES
            })


@app.route('/api/retry-all', methods=['POST'])
def retry_all_failed():
    """Retry all failed websites"""
    global monitoring_results

    with results_lock:
        failed_sites = [f.copy() for f in monitoring_results['failed']]
        monitoring_results['retry_in_progress'] = True

    if not failed_sites:
        with results_lock:
            monitoring_results['retry_in_progress'] = False
        return jsonify({'success': True, 'message': 'No failed sites', 'results': []})

    results = []

    for site in failed_sites:
        retry_count = site.get('retry_count', 0)

        if retry_count >= MAX_RETRIES:
            results.append({'url': site['url'], 'skipped': True, 'reason': 'Max retries'})
            continue

        result = check_website(site)

        with results_lock:
            if result['success']:
                monitoring_results['failed'] = [f for f in monitoring_results['failed'] if f['url'] != site['url']]
                results.append({'url': site['url'], 'success': True})
            else:
                for f in monitoring_results['failed']:
                    if f['url'] == site['url']:
                        f['retry_count'] = retry_count + 1
                        f['last_retry'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                        break
                results.append({'url': site['url'], 'success': False})

        time.sleep(0.5)

    with results_lock:
        monitoring_results['retry_in_progress'] = False

    successful = sum(1 for r in results if r.get('success'))

    return jsonify({
        'success': True,
        'total': len(failed_sites),
        'successful': successful,
        'failed': len(failed_sites) - successful,
        'remaining_failed': len(monitoring_results['failed'])
    })


if __name__ == '__main__':
    print("=" * 60)
    print("Adani Website Health Monitor")
    print("=" * 60)

    # Auto-start monitoring
    if not monitoring_results['is_running']:
        t = threading.Thread(target=monitor_websites, daemon=True)
        t.start()
        print("🚀 Auto-started monitoring")

    # Run with threading enabled
    app.run(debug=True, host='0.0.0.0', port=5000, threaded=True)